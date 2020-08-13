from typing import List

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Font, Side, Alignment
from openpyxl.formatting.rule import FormulaRule

try:
    from cue import Cue
except ModuleNotFoundError:  # This is for lambda...
    from .cue import Cue


class Nihonbashi:
    def write(self, cues: List[Cue], filepath: str):
        wb = Workbook()
        ws = wb.active
        ws.title = "Cue"

        font8 = Font(name='メイリオ', size=8)
        font11 = Font(name='メイリオ', size=11)
        font12 = Font(name='メイリオ', size=12)

        purple_fill = PatternFill(start_color='CCCCFF', end_color='CCCCFF', fill_type='solid')
        black_txt = Font(color='000000')
        ws.conditional_formatting.add(f'A3:K{len(cues) + 3}', FormulaRule(
            formula=['MOD(ROW(),2)=0'], stopIfTrue=True, fill=purple_fill, font=black_txt))

        ws["A2"] = "キュー"
        ws["B2"] = "区間距離"
        ws["C2"] = "PC毎距離"
        ws["D2"] = "総距離"
        ws["E2"] = "進路"
        ws["F2"] = "交差点"
        ws["H2"] = "信号"
        ws["I2"] = "道標の方面"
        ws["J2"] = "道路"
        ws["K2"] = "ランドマーク(注意事項)"
        ws.merge_cells("F2:G2")

        for col in ws["A2":"K2"][0]:
            col.font = font8

        row = 3
        ws[f"A{row}"] = 1
        ws[f"B{row}"] = 0.0
        ws[f"B{row}"].number_format = '0.0'
        ws[f"C{row}"] = 0.0
        ws[f"C{row}"].number_format = '0.0'
        ws[f"D{row}"] = 0.0
        ws[f"D{row}"].number_format = '0.0'
        ws[f"E{row}"] = ""
        ws[f"G{row}"] = ""
        ws[f"I{row}"] = ""
        ws[f"J{row}"] = ""
        ws[f"K{row}"] = "スタート"

        for i, cue in enumerate(cues):
            row = i + 4
            ws[f"A{row}"] = f"=A{row - 1}+1"
            ws[f"B{row}"] = f"=D{row}-D{row - 1}"
            ws[f"B{row}"].number_format = '0.0'
            ws[f"C{row}"].number_format = '0.0'
            ws[f"D{row}"] = cue.distance
            ws[f"D{row}"].number_format = '0.0'
            ws[f"E{row}"] = cue.direction
            ws[f"G{row}"] = cue.intersection
            ws[f"I{row}"] = cue.sign
            ws[f"J{row}"] = cue.road
            ws[f"K{row}"] = cue.comment
            ws[f"L{row}"] = cue.src

        align1 = Alignment(horizontal="left", vertical="top", wrap_text=True, shrink_to_fit=False)
        align2 = Alignment(horizontal="center", vertical="center", wrap_text=True, shrink_to_fit=False)
        align3 = Alignment(horizontal="center", vertical="center", wrap_text=False, shrink_to_fit=True)
        align4 = Alignment(horizontal="center", vertical="center", wrap_text=False, shrink_to_fit=False)

        for row in ws["A3":f"K{len(cues) + 3}"]:
            for cell in row:
                if cell.column in (7, 9, 11):  # G, I, K
                    ws[cell.coordinate].font = font11
                else:
                    ws[cell.coordinate].font = font12

                if cell.column == 11:  # K
                    ws[cell.coordinate].alignment = align1
                elif cell.column in (9, 10):  # I, J
                    ws[cell.coordinate].alignment = align2
                elif cell.column == 7:  # G
                    ws[cell.coordinate].alignment = align3
                else:
                    ws[cell.coordinate].alignment = align4

        # Write border
        thin = Side(border_style="thin", color="000000")
        thin_border = Border(top=thin, left=thin, right=thin, bottom=thin)
        for row in ws["A2":f"K{len(cues) + 3}"]:
            for cell in row:
                ws[cell.coordinate].border = thin_border

        ws.column_dimensions["A"].width = 7
        ws.column_dimensions["B"].width = 9.5
        ws.column_dimensions["C"].width = 9.5
        ws.column_dimensions["D"].width = 9.5
        ws.column_dimensions["E"].width = 6.4
        ws.column_dimensions["F"].width = 4.5
        ws.column_dimensions["G"].width = 16
        ws.column_dimensions["H"].width = 4.5
        ws.column_dimensions["I"].width = 14
        ws.column_dimensions["J"].width = 11
        ws.column_dimensions["K"].width = 36

        wb.save(filepath)
